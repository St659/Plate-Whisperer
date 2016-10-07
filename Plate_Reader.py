
import matplotlib
import numpy as np
from scipy import stats as sci
from openpyxl import load_workbook
import openpyxl.utils as util
from copy import deepcopy
# Make sure that we are using QT5
matplotlib.use('Qt5Agg')
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt

plt.style.use(['seaborn-white', 'seaborn-notebook'])
from PyQt5 import QtWidgets, QtGui, QtCore
import sys
import string



def main():
    app = QtWidgets.QApplication(sys.argv)
    w = PlateReaderWindow()
    w.show()
    sys.exit(app.exec_())


class AbstractDataReader:

    def get_od_list(self):
        pass

    def get_model_list(self):
        pass

    def get_time(self):
        pass


class EpochDataReader(AbstractDataReader):
    def __init__(self, filename):
        workbook = load_workbook(filename)
        self.worksheet = workbook['Plate 1 - Sheet1']

        self.wellTableLayout = self.worksheet['C30:N37']
        self.model_list = list()
        self.od_list = list()

        self.set_tableview_values()
        self.results_range = 0
        self.blank_row = 0
        self.results_row =0
        self.time = list()
        self.time_rows = list()

        for row in self.worksheet.rows:

            if row[0].value == 'Blank 650':
                self.blank_row = row[0].row

            elif row[0].value == 'Results':
                self.results_row = row[0].row

        self.results_range = self.results_row - self.blank_row
        self.get_time_column()
        self.parse_time()
        self.get_data_values()

    def parse_time(self):
        for row in self.worksheet.iter_rows(min_row = int(self.time_row+1), max_row=int(self.results_row), min_col = self.time_column):
            if row[0].value and ':' in str(row[0].value):
                t_string = str(row[0].value)
                (h,m,s) =t_string.split(':')
                time_dec = float(h) + float(m)/60
                self.time.append(round(time_dec,1))
                self.time_rows.append(row[0].row)
            elif row[0].value:
                time_dec = float(row[0].value)*24
                self.time.append(round(time_dec, 1))
                self.time_rows.append(row[0].row)
        print(self.time)

    def get_data_values(self):
        plate_plan_dict = {'A': '0', 'B': '1', 'C': '2', 'D': '3', 'E': '4', 'F': '5', 'G': '6', 'H': '7'}
        # Get the well number from the table
        for col in self.worksheet.iter_cols(min_row=min(self.time_rows) - 1, max_row=max(self.time_rows),
                                            min_col=self.time_column + 1):
            well_address = col[0].value
            parsed_data = list()
            for cell in col[1:]:

                if cell.value or cell.value == 0:
                    parsed_data.append(float(cell.value))
                    row_index = int(plate_plan_dict[well_address[0]])
                    column_index = int(well_address[1:]) - 1

            print(len(parsed_data))
            self.od_list[row_index][column_index] = parsed_data



    def get_time_column(self):
        for row in self.worksheet.iter_rows(row_offset=self.blank_row):
            for cell in row:
                if cell.value == 'Time':

                    self.time_column = util.column_index_from_string(cell.column)
                    self.time_row = cell.row
                    return

    def set_tableview_values(self):

        for row in self.wellTableLayout:
            row_list = list()
            blank_data_list = list()
            for cell in row:
                if cell.value:
                    if 'SPL' in cell.value:
                        row_list.append(" ")

                    elif 'BLK' in cell.value:
                        row_list.append("B")
                else:
                    row_list.append("x")

                blank_data_list.append(" ")
            self.model_list.append(row_list)
            self.od_list.append(blank_data_list)

    def get_time(self):
        return self.time

    def get_od_list(self):
        return self.od_list

    def get_model_list(self):
        return self.model_list

class PlateReaderExcelData:
    def __init__(self, index, sample, od):
        self.index = index
        self.sample = sample
        self.od = od




class OmegaDataReader(AbstractDataReader):
    def __init__(self, filename):
        workbook = load_workbook(filename, read_only=True)

        worksheet = workbook['All Intervals']

        data_rows = list()

        self.time = list()
        time_columns = list()
        for row in worksheet.rows:
            if row[2].value == 'Time':
                for cell in row[3:]:
                    time_string = str(cell.value).split()
                    if len(time_string) == 2:
                        current_time = float(time_string[0])

                    elif len(time_string) == 4:
                        current_time = float(time_string[0]) + float(time_string[2]) / 60
                    else:
                        current_time = None

                    if current_time is not None and current_time not in self.time:
                        self.time.append(current_time)
                        time_columns.append(cell.column)

                data_range = max(time_columns) - min(time_columns)


            if len(str(row[0].value)) == 1:
                data_rows.append(row)

        data_min = 4 + data_range
        data_max = 5 + data_range * 2

        plate_reader_class_list = list()
        for row in data_rows:
            blank_corrected_data = list()
            plate_index = [str(row[0].value), str(row[1].value)]
            sample = str(row[2].value).split()[1]

            data_raw = row[data_min:data_max]
            for cell in data_raw:
                blank_corrected_data.append(cell.value)
            plate_reader_class_list.append(PlateReaderExcelData(plate_index, sample, blank_corrected_data))

        self.model_list = list()
        self.od_list = list()
        for x in range(0, 8):
            column = list()
            od_column = list()
            for y in range(0, 12):
                column.append("x")
                od_column.append(" ")
            self.model_list.append(column)
            self.od_list.append(od_column)

        plate_plan_dict = {'A': '0', 'B': '1', 'C': '2', 'D': '3', 'E': '4', 'F': '5', 'G': '6', 'H': '7'}
        for data in plate_reader_class_list:
            if data.sample == 'B':
                print(data.index[1])
                self.model_list[int(plate_plan_dict[data.index[0]])][int(data.index[1])-1] = "B"
            else:
                print(data.index[1])
                self.model_list[int(plate_plan_dict[data.index[0]])][int(data.index[1])-1] = " "
                self.od_list[int(plate_plan_dict[data.index[0]])][int(data.index[1])-1] = data.od

    def get_od_list(self):
        return self.od_list

    def get_model_list(self):
        return self.model_list

    def get_time(self):
        return self.time


class MyMplCanvas(FigureCanvas):
    """Ultimately, this is a QWidget (as well as a FigureCanvasAgg, etc.)."""

    def __init__(self, parent=None, width=12, height=10, dpi=100):
        self.press = None
        self.prevX = 0
        self.ydata = list()
        self.xdata = list()
        self.fig = Figure(figsize=(20, 1), dpi=dpi)

        self.axes = self.fig.add_subplot(111)




        self.axes.hold(True)

        #self.compute_initial_figure()

        #
        FigureCanvas.__init__(self, self.fig)
        self.setParent(parent)

        FigureCanvas.setSizePolicy(self,
                                   QtWidgets.QSizePolicy.Expanding,
                                   QtWidgets.QSizePolicy.Expanding)
        self.setMinimumWidth(600)
        FigureCanvas.updateGeometry(self)

        FigureCanvas.mpl_connect(self,'button_press_event', self.onclick)
        FigureCanvas.mpl_connect(self, 'button_release_event', self.onrelease)
        FigureCanvas.mpl_connect(self, 'motion_notify_event', self.onmove)
        FigureCanvas.mpl_connect(self, 'scroll_event', self.zoom)
        FigureCanvas.mpl_connect(self, 'pick_event', self.onpick)
        FigureCanvas.mpl_connect(self, 'figure_enter_event', self.on_enter)
        FigureCanvas.mpl_connect(self, 'figure_exit_event', self.on_exit)

    def compute_initial_figure(self):
        pass

    def onclick(self,event):
        pass

    def onpick(self,event):
        pass

    def onmove(self,event):
        pass

    def onrelease(self, event):
        pass

    def scroll(self,event):
        pass

    def on_enter(self,event):
        pass

    def on_exit(self,event):
        pass

    def zoom(self,event):
        pass

class PlateReaderGraphLine():
    def __init__(self, linenum, line_data, mean, std, growth_rate_mean, growth_rate_std, line_col, line_style, marker, legend, graph_type):
        self.line_num = linenum
        self.line_data = line_data
        self.mean = mean
        self.std = std
        self.growth_rate_mean = growth_rate_mean
        self.growth_rate_std = growth_rate_std
        self.line_col = line_col
        self.line_style = line_style
        self.marker = marker
        self.legend = legend
        self.graph_type = graph_type


class PlateReaderGraph(MyMplCanvas):
    """Simple canvas with a sine plot."""
    def __init__(self,  *args, **kwargs):
        MyMplCanvas.__init__(self, *args, **kwargs)

        self.data_line_list = list()
        self.data_line_num = 0
        self.text_list= list()


        self.harmonics = []
        self.overline = False
        self.active = True
        self.legend = None
        self.sub_graph = False


    def plot_figure(self,time,data, line_num, line_col, line_style, marker, legend, error_type, graph_type, sub_graph):
        self.sub_graph = sub_graph
        self.time = time
        legend_pos = False
        growth_rates = list()
        for index, line in enumerate(self.data_line_list):
            if line.line_num == line_num:
                self.data_line_list.pop(index)

        try:
            legend_pos = self.legend._loc
        except AttributeError:
            print("Legend Not Set Yet")


        data_array = np.asarray(data)

        for line in data_array:
            growth_rates.append(np.gradient(line))
        mean_array = np.mean(data_array, axis =0)
        std_array = self.calculate_error(data_array,error_type)
        growth_rate_mean = np.mean(np.asarray(growth_rates),axis=0)
        growth_rate_std = self.calculate_error(np.asarray(growth_rates), error_type)
        print(growth_rate_mean)
        print(growth_rate_std)
        self.data_line_list.insert(line_num,PlateReaderGraphLine(line_num, data_array,mean_array, std_array, growth_rate_mean, growth_rate_std, line_col, line_style, marker, legend, graph_type))
        self.fig.clf()
        legends = list()
        for line in self.data_line_list:
            self.set_plot(line)
            legends.append(line.legend)
        self.update_legend(legends, legend_pos)
        self.draw()

    def clear_line(self, line_num):
        self.data_line_list.pop(line_num)
        legend_pos = self.legend._loc
        self.fig.clf()
        legends = list()
        for line in self.data_line_list:
            self.set_plot(line)
            legends.append(line.legend)
        self.update_legend(legends, legend_pos)
        self.draw()

    def set_line_colour(self, line_num, line_col):
        legend_pos = self.legend._loc
        self.fig.clf()
        legends = list()
        for line in self.data_line_list:
            if int(line.line_num) == int(line_num):
                line.line_col = line_col
            legends.append(line.legend)
            self.set_plot(line)
        self.update_legend(legends, legend_pos)
        self.draw()

    def set_line_style(self, line_num, line_style):
        legend_pos = self.legend._loc
        self.fig.clf()
        legends = list()
        for line in self.data_line_list:
            if int(line.line_num) == int(line_num):
                line.line_style = line_style
            legends.append(line.legend)
            self.set_plot(line)
        self.update_legend(legends, legend_pos)
        self.draw()

    def set_line_marker(self, line_num, marker):
        legend_pos = self.legend._loc
        self.fig.clf()
        legends = list()
        for line in self.data_line_list:
            if int(line.line_num) == int(line_num):
                line.marker = marker
            legends.append(line.legend)
            self.set_plot(line)

        self.update_legend(legends, legend_pos)
        self.draw()

    def set_line_legend(self, line_num, legend):
        legend_pos = self.legend._loc
        self.fig.clf()
        legends =list()
        for line in self.data_line_list:
            if int(line.line_num) == int(line_num):
                line.legend = legend
            legends.append(line.legend)
            self.set_plot(line)
        self.update_legend(legends, legend_pos)
        self.draw()

    def clear_figure(self):
        self.fig.clf()
        self.data_line_list= list()

        self.data_line_num = 0
        self.text_list = list()

    def set_error(self, error_type):
        legend_pos = self.legend._loc

        self.fig.clf()
        legends = list()
        for line in self.data_line_list:
            line.std = self.calculate_error(line.line_data, error_type)
            legends.append(line.legend)
            self.set_plot(line)
        self.update_legend(legends, legend_pos)
        self.draw()

    def calculate_error(self, data, error_type):
        if error_type == 0:
            error = None
        elif error_type == 1:
            error = np.std(data, axis=0)
        elif error_type == 2:
            error = sci.sem(data)
        return error

    def set_graph_type(self, graph_type):
        legend_pos = self.legend._loc
        self.fig.clf()
        legends = list()
        for line in self.data_line_list:
            line.graph_type = graph_type
            self.set_plot(line)
            legends.append(line.legend)
        self.update_legend(legends, legend_pos)
        self.draw()

    def get_current_plot(self):
        return self.axes

    def update_legend(self, legends, legend_pos):
        self.legend = self.axes.legend(legends)
        self.legend.draggable()
        if legend_pos:
            self.legend._loc = legend_pos
    def set_plot(self, line):
        
        if self.sub_graph:
            self.axes = self.fig.add_subplot(211)
            self.axes2 = self.fig.add_subplot(212)
            (plotline2, capline2, barlinecols) = self.axes2.errorbar(self.time, line.growth_rate_mean, line.growth_rate_std, color=line.line_col,
                                linestyle=line.line_style,
                                marker=line.marker, ecolor='0.75', capsize=3)
            for cap in capline2:
                cap.set_markeredgewidth(1)
        else:
            self.axes = self.fig.add_subplot(111)

        if line.graph_type == 0:
            (plotline,capline,barlinecols) = self.axes.errorbar(self.time, line.mean, line.std, color=line.line_col, linestyle=line.line_style,
                                          marker=line.marker, ecolor='0.75', capsize=3)
            print("OD")
        elif line.graph_type ==1:
            (plotline, capline, barlinecols) = self.axes.errorbar(self.time, line.growth_rate_mean, line.growth_rate_std, color=line.line_col,
                                          linestyle=line.line_style,
                                          marker=line.marker, ecolor='0.75',capsize=3)





        for cap in capline:
            cap.set_markeredgewidth(1)
        print("Growth Rate")


    def save(self, filedialogoutput, dpi):
        filename = filedialogoutput[0] + filedialogoutput[1]
        filetype = filedialogoutput[1][1:]
        self.fig.savefig(filename, format=filetype, dpi=dpi)

    def set_style(self, style):
        plt.style.use(['seaborn-white', style])
        legend_pos = self.legend._loc
        self.fig.clf()
        legends = list()
        for line in self.data_line_list:
            self.set_plot(line)
        self.update_legend(legends, legend_pos)
        self.draw()

    def set_subgraph(self, subgraph):
        self.sub_graph = subgraph
        legend_pos = self.legend._loc
        legends = list()
        for line in self.data_line_list:
            self.set_plot(line)
            legends.append(line.legend)
        self.update_legend(legends, legend_pos)
        self.draw()



class PlateReaderData:
    def __init__(self):
        row = list()
        selected_row = list()
        table_whitespace = list()
        data= list()


        for x in range(0, 7):
            row = list()
            data_row = list()
            for x in range(0,12):
                row.append("")
                data_row.append(np.random.random([20]))

            table_whitespace.append(row)
            data.append(data_row)
        time = np.linspace(0,20,20)
        self.table_data = [table_whitespace,time,data]

    def get_table_data(self):
        return self.table_data


class PlateReaderGraphEditor(QtWidgets.QWidget):

    def __init__(self, graph):
        super(PlateReaderGraphEditor,self).__init__()
        self.graph = graph

        legend_font_size = QtWidgets.QLabel("Legend Font Size")
        self.style_dict = {'Paper':'seaborn-paper', 'Talk':'seaborn-talk', 'Notebook':'seaborn-notebook', 'Poster':'seaborn-poster'}

        self.graph_editor_form = QtWidgets.QFormLayout()

        self.set_graph_type_form()
        self.set_graph_form()
        self.set_legend_form()
        self.set_style_form()
        self.set_error_form()
        self.set_save_form()

        main_vbox = QtWidgets.QVBoxLayout()
        main_vbox.addLayout(self.graph_editor_form)


        self.setLayout(main_vbox)

    def set_graph_type_form(self):

        graph_type_box = QtWidgets.QGroupBox("Graph Type")
        graph_type_form = QtWidgets.QFormLayout()
        self.type_button_group = QtWidgets.QButtonGroup()
        self.type_button_group.buttonClicked[int].connect(self.set_type)

        self.subgraph_check = QtWidgets.QCheckBox('Subgraph')
        self.subgraph_check.stateChanged.connect(self.set_subgraph)

        od_button = QtWidgets.QRadioButton("OD")
        growth_rate_button = QtWidgets.QRadioButton("Growth Rate")
        od_button.setChecked(True)

        self.type_button_group.addButton(od_button, 0)
        self.type_button_group.addButton(growth_rate_button, 1)

        type_hbox = QtWidgets.QHBoxLayout()
        type_hbox.addWidget(od_button)
        type_hbox.addWidget(growth_rate_button)

        graph_type_form.addRow(type_hbox)
        graph_type_form.addRow(self.subgraph_check)
        graph_type_box.setLayout(graph_type_form)

        self.graph_editor_form.addRow(graph_type_box)

    def set_legend_form(self):
        graph_legend_form = QtWidgets.QFormLayout()
        graph_legend_box = QtWidgets.QGroupBox("Legend")

        self.legend_fontsize_box = QtWidgets.QComboBox()
        self.legend_fontsize_box.addItems(['6','8','10','12','14'])

        graph_legend_form.addRow('Font Size', self.legend_fontsize_box)
        graph_legend_box.setLayout(graph_legend_form)
        self.graph_editor_form.addRow(graph_legend_box)

    def set_style_form(self):
        self.graph_style_form = QtWidgets.QFormLayout()
        self.graph_style_box = QtWidgets.QGroupBox("Style")

        self.graph_style_cbox = QtWidgets.QComboBox()
        self.graph_style_cbox.addItems(['Paper', 'Talk', 'Notebook', 'Poster'])
        self.graph_style_cbox.currentIndexChanged.connect(self.set_style)

        self.graph_style_form.addRow('Graph Style', self.graph_style_cbox)
        self.graph_style_box.setLayout(self.graph_style_form)
        self.graph_editor_form.addRow(self.graph_style_box)

    def set_graph_form(self):
        self.title_edit = QtWidgets.QLineEdit()
        self.xaxis_edit = QtWidgets.QLineEdit()
        self.yaxis_edit = QtWidgets.QLineEdit()


        self.title_edit.textChanged.connect(self.set_title)
        self.xaxis_edit.textChanged.connect(self.set_xaxis)
        self.yaxis_edit.textChanged.connect(self.set_yaxis)


        self.graph_labels_form = QtWidgets.QFormLayout()
        self.graph_labels_box = QtWidgets.QGroupBox("Labels")

        self.graph_labels_form.addRow('Title', self.title_edit)
        self.graph_labels_form.addRow('X Axis', self.xaxis_edit)
        self.graph_labels_form.addRow('Y Axis', self.yaxis_edit)
        self.graph_labels_box.setLayout(self.graph_labels_form)
        self.graph_editor_form.addRow(self.graph_labels_box)

    def set_error_form(self):
        self.error_button_group = QtWidgets.QButtonGroup()
        self.error_button_group.buttonClicked[int].connect(self.set_error)
        noerror_button = QtWidgets.QRadioButton("None")
        stdev_button = QtWidgets.QRadioButton("Std Dev")
        stderr_button = QtWidgets.QRadioButton("Std Error")
        stdev_button.setChecked(True)

        self.error_button_group.addButton(noerror_button, 0)
        self.error_button_group.addButton(stdev_button, 1)
        self.error_button_group.addButton(stderr_button, 2)

        self.error_button_groupbox = QtWidgets.QGroupBox("Error")

        self.error_hbox = QtWidgets.QHBoxLayout()
        self.error_hbox.addWidget(noerror_button)
        self.error_hbox.addWidget(stdev_button)
        self.error_hbox.addWidget(stderr_button)


        self.error_button_groupbox.setLayout(self.error_hbox)
        self.graph_editor_form.addRow(self.error_button_groupbox)

    def set_save_form(self):
        self.graph_save_form = QtWidgets.QFormLayout()
        self.graph_save_group = QtWidgets.QGroupBox('Save')
        self.save_button = QtWidgets.QPushButton('Save Graph')
        self.save_button.clicked.connect(self.save_graph)
        self.dpi_edit = QtWidgets.QLineEdit('100')
        self.filetype_box = QtWidgets.QComboBox()

        self.graph_save_form.addRow('DPI', self.dpi_edit)


        self.graph_save_form.addRow(self.save_button)

        self.graph_save_group.setLayout(self.graph_save_form)
        self.graph_editor_form.addRow(self.graph_save_group)

    def set_subgraph(self, state):
        self.graph.set_subgraph(self.subgraph_check.isChecked())
        self.redraw_titles()

    def save_graph(self):
        print("Graph Saved")
        file_dialog = QtWidgets.QFileDialog.getSaveFileName(self, 'Save file', '.', ".png;;.pdf" )
        dpi = int(self.dpi_edit.text())

        if dpi > 80 and dpi <2000:
            self.graph.save(file_dialog, dpi)
        else:
            print("Please Select a DPI between 80 and 2000")

    def set_title(self):
        new_text = self.title_edit.text()
        self.graph.get_current_plot().set_title(new_text)
        self.graph.draw()

    def set_xaxis(self):
        new_text = self.xaxis_edit.text()
        try:
            self.graph.get_current_plot().set_xlabel(new_text)
        except ValueError:
            print("Having Some Problems With Text")
        self.graph.draw()

    def set_style(self):
        style = self.graph_style_cbox.currentText()
        self.graph.set_style(self.style_dict[style])
        self.redraw_titles()

    def set_yaxis(self):
        new_text = self.yaxis_edit.text()
        try:
            self.graph.get_current_plot().set_ylabel(new_text)
        except ValueError:
            print("Having Some Problems With Text")
        self.graph.draw()

    def set_error(self, selected_id):
        self.graph.set_error(selected_id)
        self.redraw_titles()

    def set_type(self, selected_id):
        print(selected_id)
        self.graph.set_graph_type(selected_id)
        self.redraw_titles()

    def get_checked_error(self):
        return self.error_button_group.checkedId()

    def get_checked_type(self):
        return self.type_button_group.checkedId()

    def redraw_titles(self):
        self.set_title()
        self.set_xaxis()
        self.set_yaxis()

    def get_subgraph(self):
        return self.subgraph_check.isChecked()

class PlateReaderLine(QtWidgets.QWidget):
    def __init__(self, line_num, graph, graph_editor):
        super(PlateReaderLine, self).__init__()
        self.line_num = line_num
        self.graph = graph
        self.graph_editor = graph_editor
        self.colour_dict = {'Blue':'b', 'Red': 'r', 'Green': 'g', 'Black':'k'}
        self.marker_dict = {'None':'None', 'Square': "s", 'Circle':'o', 'Triangle':'^'}


        self.legend_edit = QtWidgets.QLineEdit(str(line_num))
        self.legend_edit.textChanged.connect(self.set_legend)
        self.num_label = QtWidgets.QLabel(str(line_num))
        line_col_label = QtWidgets.QLabel("Line Colour")
        line_style_label = QtWidgets.QLabel("Line Style")
        line_marker_label = QtWidgets.QLabel("Marker")

        self.line_col_box = QtWidgets.QComboBox()
        self.line_col_box.addItems(['Blue', 'Green', 'Red', 'Black'])
        self.line_col_box.currentIndexChanged.connect(self.set_col)

        self.line_style_box = QtWidgets.QComboBox()
        self.line_style_box.addItems(['-', '--','-.',':'])
        self.line_style_box.currentIndexChanged.connect(self.set_style)

        self.line_marker_box = QtWidgets.QComboBox()
        self.line_marker_box.addItems(['None', 'Square', 'Circle', 'Triangle'])
        self.line_marker_box.currentIndexChanged.connect(self.set_marker)

        hbox = QtWidgets.QHBoxLayout()
        hbox.setSizeConstraint(QtWidgets.QLayout.SetFixedSize)
        hbox.addWidget(self.legend_edit)
        hbox.addWidget(self.num_label)
        hbox.addWidget(line_col_label)
        hbox.addWidget(self.line_col_box)
        hbox.addWidget(line_style_label)
        hbox.addWidget(self.line_style_box)
        hbox.addWidget(line_marker_label)
        hbox.addWidget(self.line_marker_box)

        self.setLayout(hbox)
        self.setMinimumWidth(600)
    def get_line_num(self):
        return self.line_num

    def get_line_col(self):
        colour = self.line_col_box.currentText()
        return self.colour_dict[colour]

    def get_line_style(self):
        return self.line_style_box.currentText()

    def get_line_marker(self):
        marker = self.line_marker_box.currentText()
        return self.marker_dict[marker]

    def get_line_legend(self):
        legend = self.legend_edit.text()
        return legend

    def set_marker(self):
        marker = self.line_marker_box.currentText()
        marker_val = self.marker_dict[marker]
        self.graph.set_line_marker(self.get_line_num(), marker_val)
        self.graph_editor.redraw_titles()

    def set_col(self):
        colour = self.line_col_box.currentText()
        colour_val = self.colour_dict[colour]
        self.graph.set_line_colour(self.get_line_num(),colour_val)
        self.graph_editor.redraw_titles()

    def set_style(self):
        style = self.line_style_box.currentText()
        self.graph.set_line_style(self.get_line_num(), style)
        self.graph_editor.redraw_titles()

    def set_legend(self):
        legend = self.legend_edit.text()
        self.graph.set_line_legend(self.get_line_num(),legend)
        self.graph_editor.redraw_titles()

class PlateTableButtons(QtWidgets.QWidget):
    def __init__(self, tableview, tablemodel, graph, graph_editor):
        super(PlateTableButtons, self).__init__()
        self.tableview = tableview
        self.tablemodel = tablemodel
        self.graph = graph
        self.graph_editor = graph_editor
        self.line_count = 0
        self.current_line_indexes = list()

        button_box = QtWidgets.QVBoxLayout()
        add_line_button = QtWidgets.QPushButton("Add Line")
        add_line_button.clicked.connect(self.add_line)
        set_data_line_button = QtWidgets.QPushButton("Set Data")
        set_data_line_button.clicked.connect(self.set_data_line)
        clear_cells_button = QtWidgets.QPushButton("Clear Line")
        clear_cells_button.clicked.connect(self.clear_line)

        self.line_list = QtWidgets.QListWidget()
        self.line_list.itemClicked.connect(self.line_selected)
        self.line_list.setMinimumWidth(600)

        button_box.addWidget(self.line_list)
        button_box.addWidget(add_line_button)
        button_box.addWidget(set_data_line_button)
        button_box.addWidget(clear_cells_button)
        self.setLayout(button_box)


    def set_data_line(self):
        indexes = self.tableview.selectionModel().selectedIndexes()
        line_data = list()
        try:
            selected_list_index = self.line_list.selectionModel().selectedIndexes()[0].row()
        except IndexError:
            print("Please select a line")
            return
        print("Selected Index: " + str(selected_list_index))
        item = self.line_list.item(selected_list_index)
        line_item = self.line_list.itemWidget(item)
        self.clear_cells(selected_list_index)
        for index in indexes:
            self.tablemodel.setData(index, line_item.get_line_num())
            self.tableview.selectionModel().select(index, QtCore.QItemSelectionModel.Deselect)
            line_data.append(self.tablemodel.get_graph_data(index))
        self.current_line_indexes[selected_list_index] = indexes

        print(line_data)
        valid_line_data = list()
        try:
            invalid = 0
            for index, data in enumerate(line_data):
                print(index)
                if data is not None:
                    valid_line_data.append(data)
        except TypeError:
            print("Please select some data points!")
            return
        if valid_line_data:
            self.graph.plot_figure(self.tablemodel.time, valid_line_data, int(line_item.get_line_num()),
                                   line_item.get_line_col(),
                                   line_item.get_line_style(), line_item.get_line_marker(),
                                   line_item.get_line_legend(),
                                   self.graph_editor.get_checked_error(), self.graph_editor.get_checked_type(),
                                   self.graph_editor.get_subgraph())
            self.graph_editor.redraw_titles()

    def add_line(self):
        self.line_count += 1
        line = PlateReaderLine(self.line_count, self.graph, self.graph_editor)
        item = QtWidgets.QListWidgetItem()
        item.setSizeHint(line.sizeHint())
        self.line_list.addItem(item)
        self.line_list.setItemWidget(item, line)
        self.line_list.item(self.line_count - 1).setSelected(True)
        self.line_list.setFocus()
        self.current_line_indexes.append([])

    def clear_line(self):
        indexes = self.tableview.selectionModel().selectedIndexes()
        selected_list_index = self.line_list.selectionModel().selectedIndexes()[0].row()
        self.clear_cells(selected_list_index)
        for index in indexes:
            self.tablemodel.setData(index, "")
            self.tableview.selectionModel().select(index, QtCore.QItemSelectionModel.Deselect)
        self.graph.clear_line(selected_list_index)

    def clear_cells(self, index):
        current_line = self.current_line_indexes[index]
        if not current_line:
            print(self.current_line_indexes[index])
            pass
        else:
            for index in current_line:
                self.tablemodel.setData(index, "")

    def line_selected(self):
        print(len(self.line_list.selectedItems()))
        self.tableview.set_current_line(self.line_list.selectedItems()[0].text())

class PlateTable(QtWidgets.QWidget):
    def __init__(self, tablemodel, graph, graph_editor):
        super(PlateTable, self).__init__()
        self.tablemodel = tablemodel
        self.graph = graph
        self.graph_editor = graph_editor

        v_box = QtWidgets.QVBoxLayout()
        self.tableview = PlateReaderTableView(self.tablemodel)
        buttons = PlateTableButtons(self.tableview, self.tablemodel, self.graph, self.graph_editor)

        v_box.addWidget(self.tableview)
        v_box.addWidget(buttons)
        self.setLayout(v_box)

class PlateGraphTab(QtWidgets.QWidget):
    def __init__(self, tablemodel):
        super(PlateGraphTab, self).__init__()
        hbox = QtWidgets.QHBoxLayout()


        graph = PlateReaderGraph()


        graph_editor = PlateReaderGraphEditor(graph)
        plate_table = PlateTable(tablemodel, graph, graph_editor)

        hbox.addWidget(plate_table)
        hbox.addWidget(graph)
        hbox.addWidget(graph_editor)
        self.setLayout(hbox)


class PlateReaderWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(PlateReaderWindow, self).__init__()
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)
        self.setWindowTitle("Plate Reader Graph Maker")

        self.file_menu = QtWidgets.QMenu('&File', self)
        self.file_menu.addAction('Open', self.file_open)
        self.file_menu.addAction('Add Graph', self.file_add_graph)
        self.file_menu.addAction('&Quit', self.file_quit,
                                 QtCore.Qt.CTRL + QtCore.Qt.Key_Q)
        self.menuBar().addMenu(self.file_menu)

        self.main_widget = PlateReaderMainWidget()
        self.setCentralWidget(self.main_widget)
        self.setGeometry(100,100, 1000, 600)

    def file_open(self):
        self.main_widget.open_platereader_data()

    def file_add_graph(self):
        self.main_widget.add_graph()

    def file_quit(self):
        pass

class PlateReaderMainWidget(QtWidgets.QWidget):
    def __init__(self, *args):
        super(PlateReaderMainWidget,self).__init__()
        # create table
        initial_table = self.initalise_table()
        self.tabledata = False
        self.graph_num = 1

        self.header = self.createTableHeader()
        self.tablemodel = PlateReaderTableModel(initial_table, self.header)

        # layout
        layout = QtWidgets.QVBoxLayout()
        main_box =QtWidgets.QHBoxLayout()

        self.setGeometry(100,100,1200,600)

        main_box.addLayout(layout)

        graph_tab = PlateGraphTab(self.tablemodel)
        self.tab = QtWidgets.QTabWidget()

        self.tab.addTab(graph_tab, "Graph 1")
        main_box.addWidget(self.tab)

        self.setLayout(main_box)

    def initalise_table(self):
        tableview_list = list()
        for x in range(0, 8):
            column = list()
            for y in range(0, 12):
                column.append(" ")

            tableview_list.append(column)

        return [tableview_list, [],[]]

    def add_graph(self):
        if self.tabledata:
            self.graph_num +=1
            tab_label = "Graph " + str(self.graph_num)
            table_model = PlateReaderTableModel(deepcopy(self.tabledata), self.header)
            graph_tab = PlateGraphTab(table_model)
            self.tab.addTab(graph_tab, tab_label)
        else:
            print("Please select your data first")

    def open_platereader_data(self):
        filename = QtWidgets.QFileDialog.getOpenFileName(self,'Select File', '.')

        try:
            data_reader = EpochDataReader(filename[0])
            print("Epoch Reader!")
        except KeyError:
            data_reader = OmegaDataReader(filename[0])
            print("Omega Reader!")

        model_view = data_reader.get_model_list()
        plate_data = data_reader.get_od_list()
        time = data_reader.get_time()
        self.tabledata = [model_view, time, plate_data]
        self.tablemodel.reset_data(deepcopy(self.tabledata))

    def createTableHeader(self):

        header_col = list()
        # set the table model
        for x in range(1,13):
            header_col.append(str(x))
        header_row = string.ascii_uppercase[:8]
        header = [header_col,header_row]
        print(header_row)

        return header


class PlateReaderTableView(QtWidgets.QTableView):
    def __init__(self, model):
        QtWidgets.QTableView.__init__(self)
        self.setModel(model)
        self.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectItems)
        self.current_line = ""

        # set the minimum size
        self.setMinimumSize(500, 300)

        # hide grid
        self.setShowGrid(False)

        # set the font
        font = QtGui.QFont("Courier New", 10)
        self.setFont(font)

        # hide vertical header
        vh = self.verticalHeader()
        vh.setVisible(True)
        vh.setSectionResizeMode(QtWidgets.QHeaderView.Stretch)


        # set horizontal header properties
        hh = self.horizontalHeader()
        hh.setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        # hh.setStretchLastSection(True)
        #
        #selection_model.selectionChanged.connect(self.handle_selected_cells)


        # set column width to fit contents
        self.resizeColumnsToContents()

        #delegate = PlateReaderTableDelegate()
        #elf.setItemDelegate(delegate)
        #self.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)

        #self.selectionModel().selectionChanged.connect(self.selection_changed)

    def set_current_line(self, line):
        self.current_line = line


    # def mouseReleaseEvent(self, event):
    #     for index  in self.selectionModel().selectedIndexes():
    #         self.selectionModel().reset()
    #     self.setState(QtWidgets.QAbstractItemView.NoState)


class PlateReaderTableModel(QtCore.QAbstractTableModel):
    def __init__(self, datain, headerdata, parent=None, *args):
        """ datain: a list of lists
            headerdata: a list of strings
        """
        QtCore.QAbstractTableModel.__init__(self, parent, *args)
        self.arraydata = datain[0]
        self.time = datain[1]
        self.graph_data = datain[2]
        self.current_line = ""
        self.setDataCounter = 0
        self.set_cells = list()

        self.headerdata = headerdata

    def rowCount(self, parent):
        return len(self.arraydata)

    def columnCount(self, parent):
        return len(self.arraydata[0])

    def data(self, index, role):
        if not index.isValid():
            return None
        if role == QtCore.Qt.DisplayRole:
            return self.arraydata[index.row()][index.column()]
        elif role == QtCore.Qt.TextAlignmentRole:
            return QtCore.Qt.AlignCenter

    def reset_data(self, datain):
        self.beginResetModel()
        self.arraydata = datain[0]
        self.time = datain[1]
        self.graph_data = datain[2]
        self.endResetModel()

    def get_graph_data(self, index):
        if self.arraydata[index.row()][index.column()] == 'B' or self.arraydata[index.row()][index.column()] == 'x':
            return None
        else:
            try:
                return self.graph_data[index.row()][index.column()]
            except IndexError:
                return None

    def setData(self, index, value, role=QtCore.Qt.EditRole):
        self.setDataCounter +=1
        if index.isValid():
            if self.arraydata[index.row()][index.column()] == 'B' or self.arraydata[index.row()][index.column()] == 'x':
                pass
            else:
                self.arraydata[index.row()][index.column()] = value
                self.dataChanged.emit(index,index,[])

        return True

    def headerData(self, col, orientation, role):
        if orientation == QtCore.Qt.Horizontal and role == QtCore.Qt.DisplayRole:
            return self.headerdata[0][col]
        elif orientation == QtCore.Qt.Vertical and role == QtCore.Qt.DisplayRole and col < len(self.headerdata[1]):
            return self.headerdata[1][col]

    def flags(self, index):
        return super(PlateReaderTableModel, self).flags(index) | QtCore.Qt.ItemIsUserCheckable

class PlateReaderTableDelegate(QtWidgets.QStyledItemDelegate):

    def __init__(self, parent = None):
        super(PlateReaderTableDelegate, self).__init__(parent)

    def paint(self, painter, option, index):

        painter.save()

        if option.state & QtWidgets.QStyle.State_Selected:
            painter.setBrush(QtGui.QBrush(QtCore.Qt.blue))
            painter.setPen(QtGui.QPen(QtCore.Qt.black))
            value = index.data(QtCore.Qt.DisplayRole)

            text = str(value)

            painter.drawRect(option.rect)
            #painter.drawText(option.rect, QtCore.Qt.AlignLeft, text)



        else:
            painter.setBrush(QtGui.QBrush(QtCore.Qt.white))
            painter.setPen(QtGui.QPen(QtCore.Qt.black))
            value = index.data(QtCore.Qt.DisplayRole)

            text = str(value)

            #painter.drawRect(option.rect)
            painter.drawText(option.rect, QtCore.Qt.AlignCenter, text)

        painter.restore()


    # def sort(self, Ncol, order):
    #     """Sort table by given column number.
    #     """
    #     self.emit(SIGNAL("layoutAboutToBeChanged()"))
    #     self.arraydata = sorted(self.arraydata, key=operator.itemgetter(Ncol))
    #     if order == Qt.DescendingOrder:
    #         self.arraydata.reverse()
    #     self.emit(SIGNAL("layoutChanged()"))


if __name__ == "__main__":
    main()